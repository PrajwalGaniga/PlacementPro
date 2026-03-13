"""
utils/excel_parser.py – Parse student Excel uploads.
Expected columns: Name, USN, Email, Branch
"""

import io
from typing import List, Dict, Any, Tuple

import pandas as pd
from fastapi import UploadFile, HTTPException

REQUIRED_COLUMNS = {"Name", "USN", "Email", "Branch"}


async def parse_student_excel(file: UploadFile) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Parse an Excel file with student data.
    Returns (students list, errors list).
    """
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    # Normalize column names (strip whitespace, title-case)
    df.columns = [str(c).strip().title() for c in df.columns]

    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Excel missing required columns: {', '.join(missing)}. "
                   f"Found: {', '.join(df.columns)}",
        )

    students: List[Dict[str, Any]] = []
    errors: List[str] = []

    for idx, row in df.iterrows():
        row_num = int(idx) + 2  # Excel row number (1-indexed + header)
        name = str(row.get("Name", "")).strip()
        usn = str(row.get("Usn", row.get("USN", ""))).strip().upper()
        email = str(row.get("Email", "")).strip().lower()
        branch = str(row.get("Branch", "")).strip()

        if not name or name == "nan":
            errors.append(f"Row {row_num}: Missing Name")
            continue
        if not usn or usn == "NAN":
            errors.append(f"Row {row_num}: Missing USN")
            continue
        if not email or email == "nan" or "@" not in email:
            errors.append(f"Row {row_num}: Invalid email '{email}'")
            continue
        if not branch or branch == "nan":
            errors.append(f"Row {row_num}: Missing Branch")
            continue

        students.append({
            "name": name,
            "usn": usn,
            "email": email,
            "branch": branch,
        })

    return students, errors


def generate_student_template() -> bytes:
    """Generate a sample Excel template for student upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["Name", "USN", "Email", "Branch"]
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 25

    # Sample rows
    sample_rows = [
        ["ANVITHA SHETTY", "01SU23CS304", "anvithashetty41@gmail.com", "Computer Science and Engineering"],
        ["ISHWARYA", "4SN23CG004", "ishwarya9448@gmail.com", "Computer Science and Design"],
    ]
    for row_data in sample_rows:
        ws.append(row_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
